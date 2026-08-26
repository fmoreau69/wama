"""
Capacité d'import : fichiers de données tabulaires — `.csv`, `.txt`, `.tsv`, `.xlsx`.

Deuxième lecteur livré, et il l'est délibérément : **un registre avec un seul lecteur ne prouve
rien**. Deux formats sans rien de commun (SQLite indexé d'un côté, fichier plat de l'autre) montrent
que le contrat `SourceReader` tient et que `load()` fonctionne sans savoir qui a lu.

Ce lecteur couvre la partie « fichier de données (.xlsx, .csv, .txt, etc.) » du périmètre visé.
Restent à écrire : LSL `.xdf`, RTMaps `.rec`, Rosbag `.ros`, dataframes `.dt` — chacun est un
fichier de plus dans ce dossier, aucune modification du moteur.
"""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from wama.common.catalog.data_types import DataType

from ..core.noms import COLONNES_TEMPS  # noqa: F401 — réexporté : consommateurs historiques
from ..core.temporal import NEAREST, SignalMeta
from . import SourceInfo, SourceReader, StreamSpec, register_reader


def _numerise(entetes: List[str], lignes: List[list]) -> List[list]:
    """Convertit en flottants les colonnes ENTIÈREMENT numériques. Les autres restent telles quelles.

    ⚠ POURQUOI, ET POURQUOI PAR COLONNE (défaut trouvé le 2026-08-24 par le test de chaîne
    complète). Un CSV ne porte aucun type : `csv.reader` rend des chaînes. Le lecteur ne
    convertissait que l'axe du temps, si bien qu'un jeu importé depuis un CSV traversait le
    référentiel, le pont et la `Vue` sans broncher, **puis levait dans le Calculator**
    (`fmean` : « must be real number, not str »). Un point d'entrée qui produit des données
    inexploitables n'en est pas un.

    ⚠ LA DÉCISION EST PRISE PAR COLONNE, JAMAIS PAR CELLULE. Convertir « quand ça marche » ferait
    qu'une même colonne se comparerait tantôt comme du texte, tantôt comme un nombre, selon les
    lignes — c'est exactement le défaut que `_num()` refuse dans `core/conditions.py`, et celui
    que la notion de SORTE suppose absent (une colonne a UNE sorte). On exige donc que **toutes**
    les valeurs présentes de la colonne soient numériques ; une seule ne l'est pas, la colonne
    reste du texte, et `sorte_de_colonne` la verra comme telle.

    Une cellule VIDE ne disqualifie pas la colonne et devient `None` : un trou est un trou, pas
    une valeur textuelle. `.xlsx` arrive déjà typé (openpyxl) — la fonction est idempotente.
    """
    if not lignes:
        return lignes

    def _vide(v) -> bool:
        return v is None or (isinstance(v, str) and not v.strip())

    numeriques = []
    for j in range(len(entetes)):
        vues = 0
        for r in lignes:
            v = r[j] if j < len(r) else None
            if _vide(v):
                continue
            if isinstance(v, bool) or not isinstance(v, (int, float)):
                try:
                    float(v)
                except (TypeError, ValueError):
                    break
            vues += 1
        else:
            numeriques.append(j if vues else -1)   # colonne vide : rien à convertir
    a_convertir = {j for j in numeriques if j >= 0}
    if not a_convertir:
        return lignes

    out = []
    for r in lignes:
        ligne = list(r)
        for j in a_convertir:
            if j < len(ligne):
                ligne[j] = None if _vide(ligne[j]) else float(ligne[j])
        out.append(ligne)
    return out


class TabularReader(SourceReader):
    format = 'tabular'
    extensions = ('.csv', '.txt', '.tsv', '.xlsx')
    description = "Fichier de données tabulaire (CSV/TSV/TXT/XLSX) avec une colonne temporelle"

    # ── Reconnaissance ────────────────────────────────────────────────────────────────────────
    def can_read(self, path: Path) -> bool:
        if path.suffix.lower() not in self.extensions:
            return False
        if path.suffix.lower() == '.xlsx':
            try:
                import openpyxl  # noqa: F401
            except ImportError:
                return False     # dépendance absente : on décline au lieu d'échouer à la lecture
        return True

    # ── Lecture générique ─────────────────────────────────────────────────────────────────────
    def _read_rows(self, path: Path) -> tuple:
        """Rend (en-têtes, lignes). Sépare le décodage du format de toute logique temporelle."""
        if path.suffix.lower() == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            entetes = [str(h) if h is not None else '' for h in (next(it, ()) or ())]
            lignes = [list(r) for r in it]
            wb.close()
            return entetes, lignes

        with path.open('r', encoding='utf-8-sig', errors='replace', newline='') as fh:
            echantillon = fh.read(8192)
            fh.seek(0)
            try:
                dialecte = csv.Sniffer().sniff(echantillon, delimiters=',;\t|')
            except csv.Error:
                dialecte = csv.excel        # repli : virgule, cas le plus courant
            lecteur = csv.reader(fh, dialecte)
            entetes = next(lecteur, [])
            lignes = [r for r in lecteur if r]
        return [str(h).strip() for h in entetes], lignes

    @staticmethod
    def _colonne_temps(entetes: List[str]) -> Optional[int]:
        bas = [h.strip().lower() for h in entetes]
        for candidat in COLONNES_TEMPS:
            if candidat in bas:
                return bas.index(candidat)
        return None

    # ── Inventaire ────────────────────────────────────────────────────────────────────────────
    def probe(self, path: Path) -> SourceInfo:
        entetes, lignes = self._read_rows(path)
        idx = self._colonne_temps(entetes)
        note = (f"{len(lignes)} ligne(s), {len(entetes)} colonne(s) ; "
                + (f"axe du temps = '{entetes[idx]}'" if idx is not None
                   else f"AUCUNE colonne temporelle (attendu l'un de : {', '.join(COLONNES_TEMPS)})"))
        # Un fichier plat = un seul flux, nommé d'après le fichier.
        return SourceInfo(format=self.format, path=str(path),
                          streams=[path.stem] if idx is not None else [],
                          attributes={'columns': entetes, 'rows': len(lignes)},
                          notes=note)

    def read(self, path: Path, streams: Optional[Iterable[str]] = None,
             timestampers: Optional[Dict[str, Any]] = None) -> List[StreamSpec]:
        timestampers = timestampers or {}
        entetes, lignes = self._read_rows(path)
        idx = self._colonne_temps(entetes)
        if idx is None:
            raise ValueError(
                f"'{path.name}' n'a pas de colonne temporelle reconnue "
                f"(attendu l'un de : {', '.join(COLONNES_TEMPS)})")

        name = path.stem
        if streams is not None and name not in list(streams):
            return []

        paires = []
        for r in lignes:
            try:
                paires.append((float(r[idx]), r))
            except (TypeError, ValueError, IndexError):
                continue          # ligne sans temps exploitable : écartée, pas devinée
        paires.sort(key=lambda p: p[0])
        times = [p[0] for p in paires]
        # Typage PAR COLONNE avant exposition : sans lui, un CSV traverse tout et casse au premier
        # calcul (voir `_numerise`). Fait ici, une seule fois, et non à chaque lecture de tranche.
        donnees = _numerise(entetes, [p[1] for p in paires])

        ts = timestampers.get(name)
        if ts is not None:
            times = [ts.timestamp(t, i, t) for i, t in enumerate(times)]

        def rows(i0: int, i1: int):
            return [dict(zip(entetes, r)) for r in donnees[i0:i1]]

        meta = SignalMeta(name=name, data_type=DataType.TIMESERIES, fs=None,
                          default_lookup=NEAREST,
                          comments=f"tabulaire · {len(entetes)} colonne(s)")
        return [StreamSpec(meta=meta, times=times, rows=rows)]


register_reader(TabularReader())
